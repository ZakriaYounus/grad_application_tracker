"""
Grad/Scholarship Application Tracker
A local-first dashboard for managing graduate school and scholarship applications:
per-application checklists, recommender/LOR tracking, document version history,
and deadline alerts (with .ics calendar export).

Run with:  python app.py
Then open: http://localhost:5001
"""

import os
import sqlite3
import uuid
from datetime import datetime, date
from flask import Flask, request, jsonify, send_from_directory, send_file, g
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "tracker.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {
    "pdf", "doc", "docx", "txt", "rtf", "odt",
    "jpg", "jpeg", "png", "webp",
    "ppt", "pptx", "zip", "key"
}

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024  # 25MB per upload

DEFAULT_CHECKLIST = [
    ("Statement of Purpose", "SOP"),
    ("CV / Resume", "CV"),
    ("Official Transcripts", "Transcript"),
    ("English Test (IELTS/TOEFL)", "Test Score"),
    ("Portfolio / Writing Sample", "Portfolio"),
    ("Online Application Form", "Form"),
    ("Application Fee Payment", "Fee"),
]

STATUS_VALUES = [
    "researching", "in_progress", "submitted",
    "interview", "accepted", "rejected", "waitlisted", "declined", "withdrawn"
]


# --------------------------------------------------------------------------
# Database helpers
# --------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS applications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            university TEXT NOT NULL,
            program TEXT NOT NULL,
            degree_type TEXT DEFAULT '',
            country TEXT DEFAULT '',
            city TEXT DEFAULT '',
            funding_type TEXT DEFAULT '',
            application_deadline TEXT,
            scholarship_deadline TEXT,
            decision_date TEXT,
            status TEXT DEFAULT 'researching',
            priority TEXT DEFAULT 'target',
            application_fee REAL,
            fee_paid INTEGER DEFAULT 0,
            application_url TEXT DEFAULT '',
            portal_username TEXT DEFAULT '',
            test_required TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS checklist_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            application_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            category TEXT DEFAULT 'Other',
            status TEXT DEFAULT 'not_started',
            due_date TEXT,
            notes TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_at TEXT,
            FOREIGN KEY (application_id) REFERENCES applications(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            application_id INTEGER NOT NULL,
            category TEXT DEFAULT 'Other',
            filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            version_number INTEGER DEFAULT 1,
            is_current INTEGER DEFAULT 1,
            file_size INTEGER DEFAULT 0,
            notes TEXT DEFAULT '',
            uploaded_at TEXT,
            FOREIGN KEY (application_id) REFERENCES applications(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS recommenders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT DEFAULT '',
            institution TEXT DEFAULT '',
            title TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS lor_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            application_id INTEGER NOT NULL,
            recommender_id INTEGER NOT NULL,
            status TEXT DEFAULT 'not_requested',
            due_date TEXT,
            requested_date TEXT,
            submitted_date TEXT,
            notes TEXT DEFAULT '',
            FOREIGN KEY (application_id) REFERENCES applications(id) ON DELETE CASCADE,
            FOREIGN KEY (recommender_id) REFERENCES recommenders(id) ON DELETE CASCADE
        );
        """
    )
    db.commit()
    db.close()


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# --------------------------------------------------------------------------
# Computation helpers
# --------------------------------------------------------------------------

def days_until(date_str):
    if not date_str:
        return None
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    return (d - date.today()).days


def urgency_for(days):
    """Returns a bucket used by the frontend to color-code deadlines."""
    if days is None:
        return "none"
    if days < 0:
        return "past"
    if days <= 7:
        return "critical"
    if days <= 21:
        return "soon"
    if days <= 45:
        return "upcoming"
    return "far"


def next_deadline_info(app_row):
    """Pick the soonest non-past deadline between application & scholarship deadlines.
    Falls back to the soonest past one if both have passed (so it's still visible)."""
    candidates = []
    if app_row["application_deadline"]:
        candidates.append(("Application", app_row["application_deadline"]))
    if app_row["scholarship_deadline"]:
        candidates.append(("Scholarship", app_row["scholarship_deadline"]))
    if not candidates:
        return None
    future = [(label, d, days_until(d)) for label, d in candidates if days_until(d) is not None and days_until(d) >= 0]
    if future:
        future.sort(key=lambda x: x[2])
        label, d, days = future[0]
    else:
        allc = [(label, d, days_until(d)) for label, d in candidates]
        allc.sort(key=lambda x: x[2])
        label, d, days = allc[-1]
    return {"label": label, "date": d, "days": days, "urgency": urgency_for(days)}


def serialize_application(row, db):
    app_dict = dict(row)
    checklist = db.execute(
        "SELECT * FROM checklist_items WHERE application_id = ? ORDER BY sort_order, id",
        (row["id"],),
    ).fetchall()
    total = len(checklist)
    done = sum(1 for c in checklist if c["status"] == "done")
    lor_rows = db.execute(
        "SELECT * FROM lor_requests WHERE application_id = ?", (row["id"],)
    ).fetchall()
    lor_total = len(lor_rows)
    lor_done = sum(1 for l in lor_rows if l["status"] == "submitted")

    combined_total = total + lor_total
    combined_done = done + lor_done
    completion_pct = round((combined_done / combined_total) * 100) if combined_total else 0

    app_dict["completion_pct"] = completion_pct
    app_dict["checklist_total"] = total
    app_dict["checklist_done"] = done
    app_dict["lor_total"] = lor_total
    app_dict["lor_done"] = lor_done
    app_dict["next_deadline"] = next_deadline_info(row)
    app_dict["fee_paid"] = bool(row["fee_paid"])
    return app_dict


# --------------------------------------------------------------------------
# Static frontend
# --------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory(app.template_folder, "index.html")


# --------------------------------------------------------------------------
# Applications
# --------------------------------------------------------------------------

@app.route("/api/applications", methods=["GET"])
def list_applications():
    db = get_db()
    rows = db.execute("SELECT * FROM applications ORDER BY application_deadline IS NULL, application_deadline ASC").fetchall()
    return jsonify([serialize_application(r, db) for r in rows])


@app.route("/api/applications/<int:app_id>", methods=["GET"])
def get_application(app_id):
    db = get_db()
    row = db.execute("SELECT * FROM applications WHERE id = ?", (app_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404

    data = serialize_application(row, db)
    data["checklist"] = [dict(r) for r in db.execute(
        "SELECT * FROM checklist_items WHERE application_id = ? ORDER BY sort_order, id", (app_id,)
    ).fetchall()]

    docs = db.execute(
        "SELECT * FROM documents WHERE application_id = ? ORDER BY category, version_number DESC", (app_id,)
    ).fetchall()
    data["documents"] = [dict(r) for r in docs]

    lors = db.execute(
        """SELECT lor_requests.*, recommenders.name as recommender_name,
                  recommenders.email as recommender_email,
                  recommenders.institution as recommender_institution
           FROM lor_requests JOIN recommenders ON lor_requests.recommender_id = recommenders.id
           WHERE application_id = ? ORDER BY lor_requests.id""",
        (app_id,),
    ).fetchall()
    data["lor_requests"] = [dict(r) for r in lors]

    return jsonify(data)


@app.route("/api/applications", methods=["POST"])
def create_application():
    db = get_db()
    payload = request.get_json(force=True)
    if not payload.get("university") or not payload.get("program"):
        return jsonify({"error": "university and program are required"}), 400

    ts = now_iso()
    cur = db.execute(
        """INSERT INTO applications
           (university, program, degree_type, country, city, funding_type,
            application_deadline, scholarship_deadline, decision_date, status, priority,
            application_fee, fee_paid, application_url, portal_username, test_required,
            notes, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            payload.get("university", "").strip(),
            payload.get("program", "").strip(),
            payload.get("degree_type", ""),
            payload.get("country", ""),
            payload.get("city", ""),
            payload.get("funding_type", ""),
            payload.get("application_deadline") or None,
            payload.get("scholarship_deadline") or None,
            payload.get("decision_date") or None,
            payload.get("status", "researching"),
            payload.get("priority", "target"),
            payload.get("application_fee") or None,
            1 if payload.get("fee_paid") else 0,
            payload.get("application_url", ""),
            payload.get("portal_username", ""),
            payload.get("test_required", ""),
            payload.get("notes", ""),
            ts, ts,
        ),
    )
    new_id = cur.lastrowid

    if payload.get("use_default_checklist", True):
        for i, (title, category) in enumerate(DEFAULT_CHECKLIST):
            db.execute(
                "INSERT INTO checklist_items (application_id, title, category, status, sort_order, created_at) VALUES (?,?,?,?,?,?)",
                (new_id, title, category, "not_started", i, ts),
            )

    db.commit()
    row = db.execute("SELECT * FROM applications WHERE id = ?", (new_id,)).fetchone()
    return jsonify(serialize_application(row, db)), 201


@app.route("/api/applications/<int:app_id>", methods=["PUT"])
def update_application(app_id):
    db = get_db()
    row = db.execute("SELECT * FROM applications WHERE id = ?", (app_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    payload = request.get_json(force=True)

    fields = [
        "university", "program", "degree_type", "country", "city", "funding_type",
        "application_deadline", "scholarship_deadline", "decision_date", "status",
        "priority", "application_fee", "fee_paid", "application_url",
        "portal_username", "test_required", "notes",
    ]
    updates, values = [], []
    for f in fields:
        if f in payload:
            v = payload[f]
            if f == "fee_paid":
                v = 1 if v else 0
            if f in ("application_deadline", "scholarship_deadline", "decision_date") and v == "":
                v = None
            updates.append(f"{f} = ?")
            values.append(v)
    updates.append("updated_at = ?")
    values.append(now_iso())
    values.append(app_id)

    db.execute(f"UPDATE applications SET {', '.join(updates)} WHERE id = ?", values)
    db.commit()
    row = db.execute("SELECT * FROM applications WHERE id = ?", (app_id,)).fetchone()
    return jsonify(serialize_application(row, db))


@app.route("/api/applications/<int:app_id>", methods=["DELETE"])
def delete_application(app_id):
    db = get_db()
    docs = db.execute("SELECT * FROM documents WHERE application_id = ?", (app_id,)).fetchall()
    for d in docs:
        path = os.path.join(UPLOAD_DIR, str(app_id), d["stored_filename"])
        if os.path.exists(path):
            os.remove(path)
    app_folder = os.path.join(UPLOAD_DIR, str(app_id))
    if os.path.isdir(app_folder) and not os.listdir(app_folder):
        os.rmdir(app_folder)
    db.execute("DELETE FROM applications WHERE id = ?", (app_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Checklist items
# --------------------------------------------------------------------------

@app.route("/api/applications/<int:app_id>/checklist", methods=["POST"])
def add_checklist_item(app_id):
    db = get_db()
    payload = request.get_json(force=True)
    if not payload.get("title"):
        return jsonify({"error": "title is required"}), 400
    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order), -1) as m FROM checklist_items WHERE application_id = ?", (app_id,)
    ).fetchone()["m"]
    cur = db.execute(
        "INSERT INTO checklist_items (application_id, title, category, status, due_date, notes, sort_order, created_at) VALUES (?,?,?,?,?,?,?,?)",
        (
            app_id, payload["title"], payload.get("category", "Other"),
            payload.get("status", "not_started"), payload.get("due_date") or None,
            payload.get("notes", ""), max_order + 1, now_iso(),
        ),
    )
    db.commit()
    row = db.execute("SELECT * FROM checklist_items WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/checklist/<int:item_id>", methods=["PUT"])
def update_checklist_item(item_id):
    db = get_db()
    row = db.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    payload = request.get_json(force=True)
    fields = ["title", "category", "status", "due_date", "notes", "sort_order"]
    updates, values = [], []
    for f in fields:
        if f in payload:
            v = payload[f]
            if f == "due_date" and v == "":
                v = None
            updates.append(f"{f} = ?")
            values.append(v)
    if not updates:
        return jsonify(dict(row))
    values.append(item_id)
    db.execute(f"UPDATE checklist_items SET {', '.join(updates)} WHERE id = ?", values)
    db.commit()
    row = db.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/checklist/<int:item_id>", methods=["DELETE"])
def delete_checklist_item(item_id):
    db = get_db()
    db.execute("DELETE FROM checklist_items WHERE id = ?", (item_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Documents (version tracking)
# --------------------------------------------------------------------------

@app.route("/api/applications/<int:app_id>/documents", methods=["POST"])
def upload_document(app_id):
    db = get_db()
    app_row = db.execute("SELECT id FROM applications WHERE id = ?", (app_id,)).fetchone()
    if not app_row:
        return jsonify({"error": "application not found"}), 404

    if "file" not in request.files:
        return jsonify({"error": "no file provided"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "no file selected"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "file type not allowed"}), 400

    category = request.form.get("category", "Other")
    notes = request.form.get("notes", "")

    last_version = db.execute(
        "SELECT COALESCE(MAX(version_number), 0) as v FROM documents WHERE application_id = ? AND category = ?",
        (app_id, category),
    ).fetchone()["v"]
    new_version = last_version + 1

    original_name = secure_filename(file.filename)
    ext = original_name.rsplit(".", 1)[1].lower() if "." in original_name else ""
    stored_name = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex

    folder = os.path.join(UPLOAD_DIR, str(app_id))
    os.makedirs(folder, exist_ok=True)
    filepath = os.path.join(folder, stored_name)
    file.save(filepath)
    size = os.path.getsize(filepath)

    db.execute(
        "UPDATE documents SET is_current = 0 WHERE application_id = ? AND category = ?",
        (app_id, category),
    )
    cur = db.execute(
        """INSERT INTO documents
           (application_id, category, filename, stored_filename, version_number, is_current, file_size, notes, uploaded_at)
           VALUES (?,?,?,?,?,1,?,?,?)""",
        (app_id, category, original_name, stored_name, new_version, size, notes, now_iso()),
    )
    db.commit()
    row = db.execute("SELECT * FROM documents WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/documents/<int:doc_id>/download", methods=["GET"])
def download_document(doc_id):
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    folder = os.path.join(UPLOAD_DIR, str(row["application_id"]))
    return send_from_directory(folder, row["stored_filename"], as_attachment=True, download_name=row["filename"])


@app.route("/api/documents/<int:doc_id>/make-current", methods=["PUT"])
def make_current_document(doc_id):
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    db.execute(
        "UPDATE documents SET is_current = 0 WHERE application_id = ? AND category = ?",
        (row["application_id"], row["category"]),
    )
    db.execute("UPDATE documents SET is_current = 1 WHERE id = ?", (doc_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/documents/<int:doc_id>", methods=["DELETE"])
def delete_document(doc_id):
    db = get_db()
    row = db.execute("SELECT * FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    path = os.path.join(UPLOAD_DIR, str(row["application_id"]), row["stored_filename"])
    if os.path.exists(path):
        os.remove(path)
    db.execute("DELETE FROM documents WHERE id = ?", (doc_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Recommenders & Letters of Recommendation
# --------------------------------------------------------------------------

@app.route("/api/recommenders", methods=["GET"])
def list_recommenders():
    db = get_db()
    rows = db.execute("SELECT * FROM recommenders ORDER BY name").fetchall()
    result = []
    for r in rows:
        d = dict(r)
        usage = db.execute(
            "SELECT COUNT(*) as c FROM lor_requests WHERE recommender_id = ?", (r["id"],)
        ).fetchone()["c"]
        d["used_in"] = usage
        result.append(d)
    return jsonify(result)


@app.route("/api/recommenders", methods=["POST"])
def create_recommender():
    db = get_db()
    payload = request.get_json(force=True)
    if not payload.get("name"):
        return jsonify({"error": "name is required"}), 400
    cur = db.execute(
        "INSERT INTO recommenders (name, email, institution, title, notes, created_at) VALUES (?,?,?,?,?,?)",
        (payload["name"], payload.get("email", ""), payload.get("institution", ""),
         payload.get("title", ""), payload.get("notes", ""), now_iso()),
    )
    db.commit()
    row = db.execute("SELECT * FROM recommenders WHERE id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/recommenders/<int:rec_id>", methods=["PUT"])
def update_recommender(rec_id):
    db = get_db()
    row = db.execute("SELECT * FROM recommenders WHERE id = ?", (rec_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    payload = request.get_json(force=True)
    fields = ["name", "email", "institution", "title", "notes"]
    updates, values = [], []
    for f in fields:
        if f in payload:
            updates.append(f"{f} = ?")
            values.append(payload[f])
    if updates:
        values.append(rec_id)
        db.execute(f"UPDATE recommenders SET {', '.join(updates)} WHERE id = ?", values)
        db.commit()
    row = db.execute("SELECT * FROM recommenders WHERE id = ?", (rec_id,)).fetchone()
    return jsonify(dict(row))


@app.route("/api/recommenders/<int:rec_id>", methods=["DELETE"])
def delete_recommender(rec_id):
    db = get_db()
    db.execute("DELETE FROM recommenders WHERE id = ?", (rec_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/applications/<int:app_id>/lor", methods=["POST"])
def add_lor_request(app_id):
    db = get_db()
    payload = request.get_json(force=True)
    if not payload.get("recommender_id"):
        return jsonify({"error": "recommender_id is required"}), 400
    cur = db.execute(
        """INSERT INTO lor_requests
           (application_id, recommender_id, status, due_date, requested_date, submitted_date, notes)
           VALUES (?,?,?,?,?,?,?)""",
        (
            app_id, payload["recommender_id"], payload.get("status", "not_requested"),
            payload.get("due_date") or None, payload.get("requested_date") or None,
            payload.get("submitted_date") or None, payload.get("notes", ""),
        ),
    )
    db.commit()
    row = db.execute(
        """SELECT lor_requests.*, recommenders.name as recommender_name,
                  recommenders.email as recommender_email,
                  recommenders.institution as recommender_institution
           FROM lor_requests JOIN recommenders ON lor_requests.recommender_id = recommenders.id
           WHERE lor_requests.id = ?""",
        (cur.lastrowid,),
    ).fetchone()
    return jsonify(dict(row)), 201


@app.route("/api/lor/<int:lor_id>", methods=["PUT"])
def update_lor_request(lor_id):
    db = get_db()
    row = db.execute("SELECT * FROM lor_requests WHERE id = ?", (lor_id,)).fetchone()
    if not row:
        return jsonify({"error": "not found"}), 404
    payload = request.get_json(force=True)
    fields = ["status", "due_date", "requested_date", "submitted_date", "notes"]
    updates, values = [], []
    for f in fields:
        if f in payload:
            v = payload[f]
            if f in ("due_date", "requested_date", "submitted_date") and v == "":
                v = None
            updates.append(f"{f} = ?")
            values.append(v)
    if updates:
        values.append(lor_id)
        db.execute(f"UPDATE lor_requests SET {', '.join(updates)} WHERE id = ?", values)
        db.commit()
    row = db.execute(
        """SELECT lor_requests.*, recommenders.name as recommender_name,
                  recommenders.email as recommender_email,
                  recommenders.institution as recommender_institution
           FROM lor_requests JOIN recommenders ON lor_requests.recommender_id = recommenders.id
           WHERE lor_requests.id = ?""",
        (lor_id,),
    ).fetchone()
    return jsonify(dict(row))


@app.route("/api/lor/<int:lor_id>", methods=["DELETE"])
def delete_lor_request(lor_id):
    db = get_db()
    db.execute("DELETE FROM lor_requests WHERE id = ?", (lor_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# Dashboard stats & deadlines
# --------------------------------------------------------------------------

@app.route("/api/dashboard/stats", methods=["GET"])
def dashboard_stats():
    db = get_db()
    rows = db.execute("SELECT * FROM applications").fetchall()
    apps = [serialize_application(r, db) for r in rows]

    by_status = {}
    for a in apps:
        by_status[a["status"]] = by_status.get(a["status"], 0) + 1

    by_country = {}
    for a in apps:
        c = a["country"] or "Unspecified"
        by_country[c] = by_country.get(c, 0) + 1

    deadlines = []
    for a in apps:
        nd = a["next_deadline"]
        if nd and nd["urgency"] != "past":
            deadlines.append({
                "application_id": a["id"], "university": a["university"], "program": a["program"],
                "type": nd["label"], "date": nd["date"], "days": nd["days"], "urgency": nd["urgency"],
            })
    # also surface checklist-item due dates and LOR due dates approaching soon
    item_rows = db.execute(
        """SELECT checklist_items.*, applications.university, applications.program
           FROM checklist_items JOIN applications ON checklist_items.application_id = applications.id
           WHERE checklist_items.due_date IS NOT NULL AND checklist_items.status != 'done'"""
    ).fetchall()
    for it in item_rows:
        d = days_until(it["due_date"])
        if d is not None and d >= -3:
            deadlines.append({
                "application_id": it["application_id"], "university": it["university"], "program": it["program"],
                "type": it["title"], "date": it["due_date"], "days": d, "urgency": urgency_for(d),
            })
    lor_rows = db.execute(
        """SELECT lor_requests.*, applications.university, applications.program, recommenders.name as rec_name
           FROM lor_requests
           JOIN applications ON lor_requests.application_id = applications.id
           JOIN recommenders ON lor_requests.recommender_id = recommenders.id
           WHERE lor_requests.due_date IS NOT NULL AND lor_requests.status != 'submitted'"""
    ).fetchall()
    for lr in lor_rows:
        d = days_until(lr["due_date"])
        if d is not None and d >= -3:
            deadlines.append({
                "application_id": lr["application_id"], "university": lr["university"], "program": lr["program"],
                "type": f"LOR – {lr['rec_name']}", "date": lr["due_date"], "days": d, "urgency": urgency_for(d),
            })

    deadlines.sort(key=lambda x: x["days"])

    return jsonify({
        "total_applications": len(apps),
        "by_status": by_status,
        "by_country": by_country,
        "avg_completion": round(sum(a["completion_pct"] for a in apps) / len(apps)) if apps else 0,
        "upcoming_deadlines": deadlines[:25],
    })


# --------------------------------------------------------------------------
# Export: .ics calendar + JSON backup/restore
# --------------------------------------------------------------------------

def ics_escape(text):
    return (text or "").replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")


@app.route("/api/export/ics", methods=["GET"])
def export_ics():
    db = get_db()
    rows = db.execute("SELECT * FROM applications").fetchall()
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Grad Application Tracker//EN", "CALSCALE:GREGORIAN"]

    def add_event(uid, summary, date_str, desc=""):
        if not date_str:
            return
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            return
        dt = d.strftime("%Y%m%d")
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}@gradtracker",
            f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}",
            f"DTSTART;VALUE=DATE:{dt}",
            f"DTEND;VALUE=DATE:{dt}",
            f"SUMMARY:{ics_escape(summary)}",
            f"DESCRIPTION:{ics_escape(desc)}",
            "BEGIN:VALARM", "ACTION:DISPLAY", "DESCRIPTION:Reminder", "TRIGGER:-P3D", "END:VALARM",
            "END:VEVENT",
        ])

    for r in rows:
        add_event(f"app-{r['id']}-deadline", f"Deadline: {r['university']} – {r['program']}", r["application_deadline"], r["notes"] or "")
        add_event(f"app-{r['id']}-scholarship", f"Scholarship deadline: {r['university']} – {r['program']}", r["scholarship_deadline"], "")
        add_event(f"app-{r['id']}-decision", f"Decision expected: {r['university']} – {r['program']}", r["decision_date"], "")

    item_rows = db.execute(
        """SELECT checklist_items.*, applications.university, applications.program
           FROM checklist_items JOIN applications ON checklist_items.application_id = applications.id
           WHERE checklist_items.due_date IS NOT NULL"""
    ).fetchall()
    for it in item_rows:
        add_event(f"item-{it['id']}", f"{it['title']} due – {it['university']}", it["due_date"], it["program"])

    lor_rows = db.execute(
        """SELECT lor_requests.*, applications.university, recommenders.name as rec_name
           FROM lor_requests
           JOIN applications ON lor_requests.application_id = applications.id
           JOIN recommenders ON lor_requests.recommender_id = recommenders.id
           WHERE lor_requests.due_date IS NOT NULL"""
    ).fetchall()
    for lr in lor_rows:
        add_event(f"lor-{lr['id']}", f"LOR due – {lr['rec_name']} ({lr['university']})", lr["due_date"], "")

    lines.append("END:VCALENDAR")
    ics_content = "\r\n".join(lines)

    tmp_path = os.path.join(DATA_DIR, "deadlines.ics")
    with open(tmp_path, "w") as f:
        f.write(ics_content)
    return send_file(tmp_path, as_attachment=True, download_name="grad-application-deadlines.ics", mimetype="text/calendar")


@app.route("/api/export/json", methods=["GET"])
def export_json():
    db = get_db()
    data = {
        "applications": [dict(r) for r in db.execute("SELECT * FROM applications").fetchall()],
        "checklist_items": [dict(r) for r in db.execute("SELECT * FROM checklist_items").fetchall()],
        "recommenders": [dict(r) for r in db.execute("SELECT * FROM recommenders").fetchall()],
        "lor_requests": [dict(r) for r in db.execute("SELECT * FROM lor_requests").fetchall()],
        "documents": [dict(r) for r in db.execute("SELECT * FROM documents").fetchall()],
        "exported_at": now_iso(),
    }
    tmp_path = os.path.join(DATA_DIR, "backup.json")
    import json
    with open(tmp_path, "w") as f:
        json.dump(data, f, indent=2)
    return send_file(tmp_path, as_attachment=True, download_name="grad-tracker-backup.json", mimetype="application/json")


# --------------------------------------------------------------------------
# Import: upload an Excel research file and populate the DB automatically
# --------------------------------------------------------------------------

MONTHS_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}

IMPORT_CHECKLIST = [
    ("Statement of Purpose",         "SOP"),
    ("CV / Resume",                  "CV"),
    ("Official Transcripts",         "Transcript"),
    ("English Test (IELTS / TOEFL)", "Test Score"),
    ("Portfolio / Writing Sample",   "Portfolio"),
    ("Online Application Form",      "Application"),
    ("Application Fee Payment",      "Fee"),
]


def _parse_import_deadline(text, year=2027):
    import re
    if not text:
        return None
    t = str(text).lower()
    m = re.search(r"deadline[s]?[:\s~]*(.{0,80})", t)
    chunk = m.group(1) if m else t
    # Prefer non-EU deadline when both are listed
    non_eu = re.search(r"non[- ]eu[):]?\s*([\w\s]+\d{1,2}|\d{1,2}\s+[\w]+|[\w]+)", chunk)
    if non_eu:
        chunk = non_eu.group(1)
    for name, num in MONTHS_MAP.items():
        p = re.search(rf"\b{name}\s+(\d{{1,2}})\b", chunk)
        if p:
            return f"{year}-{num:02d}-{int(p.group(1)):02d}"
        p = re.search(rf"\b(\d{{1,2}})\s+{name}\b", chunk)
        if p:
            return f"{year}-{num:02d}-{int(p.group(1)):02d}"
        p = re.search(rf"\b(late|early|mid)?\s*{name}\b", chunk)
        if p:
            qualifier = p.group(1)
            day = 25 if qualifier == "late" else 5 if qualifier == "early" else 15
            return f"{year}-{num:02d}-{day:02d}"
    return None


def _map_priority(prob):
    if not prob:
        return "target"
    p = str(prob).lower()
    if "medium-low" in p or "low-medium" in p or ("low" in p and "medium" not in p):
        return "reach"
    return "target"


def _extract_degree(program):
    if not program:
        return "MA"
    import re
    u = str(program).upper()
    if "MFA" in u:
        return "MFA"
    if "MSC" in u or re.search(r"\bMS\b", u):
        return "MSc"
    if re.search(r"\bMA\b", u):
        return "MA"
    return "MA"


@app.route("/api/import/excel", methods=["POST"])
def import_excel():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"error": "openpyxl is not installed. Run: pip install openpyxl"}), 500

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx / .xlsm files are supported"}), 400

    import io, re
    data = f.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    db = get_db()
    existing = set(
        (r["university"].strip().lower(), r["program"].strip().lower())
        for r in db.execute("SELECT university, program FROM applications").fetchall()
    )

    ts = now_iso()
    added, skipped = 0, 0
    added_list = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        country = sheet_name.strip()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(h).strip() if h is not None else "" for h in rows[0]]

        def col(row_tuple, name):
            try:
                idx = header.index(name)
                v = row_tuple[idx]
                return str(v).strip() if v is not None else ""
            except (ValueError, IndexError):
                return ""

        for raw in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in raw):
                continue
            university  = col(raw, "Uni name")
            program     = col(raw, "Course name")
            if not university or not program:
                continue
            key = (university.lower(), program.lower())
            if key in existing:
                skipped += 1
                continue

            timeline     = col(raw, "Application timeline")
            probability  = col(raw, "Probability of getting in")
            course_url   = col(raw, "Course websites")
            tuition      = col(raw, "Tuition")
            scholarships = col(raw, "Scholarships if any")
            ielts        = col(raw, "IELTS Requirement")
            portfolio    = col(raw, "Portfolio Required")
            semester     = col(raw, "Semester offered")
            requirements = col(raw, "Application requirements")
            where_apply  = col(raw, "Where to apply")

            notes_parts = []
            if timeline:      notes_parts.append(f"Timeline: {timeline}")
            if semester:      notes_parts.append(f"Semester: {semester}")
            if requirements:  notes_parts.append(f"Requirements: {requirements}")
            if tuition:       notes_parts.append(f"Tuition: {tuition}")
            if where_apply:   notes_parts.append(f"Apply via: {where_apply}")
            if probability:   notes_parts.append(f"Admission chance: {probability}")

            cur = db.execute("""
                INSERT INTO applications
                  (university, program, degree_type, country, funding_type,
                   application_deadline, status, priority,
                   test_required, application_url, notes, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                university, program, _extract_degree(program), country,
                scholarships[:120] if scholarships else "",
                _parse_import_deadline(timeline),
                "researching", _map_priority(probability),
                ielts[:80] if ielts else "",
                course_url, "\n".join(notes_parts), ts, ts,
            ))
            app_id = cur.lastrowid

            for i, (title, category) in enumerate(IMPORT_CHECKLIST):
                db.execute("""
                    INSERT INTO checklist_items
                      (application_id, title, category, status, sort_order, created_at)
                    VALUES (?,?,?,?,?,?)
                """, (app_id, title, category, "not_started", i, ts))

            # Note portfolio requirement on the checklist item
            pt = str(portfolio).lower()
            if any(w in pt for w in ("yes", "required", "mandatory", "essential")):
                db.execute(
                    "UPDATE checklist_items SET notes=? WHERE application_id=? AND category='Portfolio'",
                    (f"Required — {portfolio}", app_id)
                )

            existing.add(key)
            added += 1
            added_list.append({"university": university, "program": program, "country": country})

    db.commit()
    return jsonify({"added": added, "skipped": skipped, "applications": added_list})


if __name__ == "__main__":
    init_db()
    print("Grad Application Tracker running at http://localhost:5001")
    app.run(host="0.0.0.0", port=5001, debug=True)
