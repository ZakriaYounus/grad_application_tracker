"""
Import your Masters research Excel file into the Grad Tracker database.

Usage:
    python import_excel.py path/to/Masters_Countries_Updated.xlsx

The app must have been run at least once first (so data/tracker.db exists).
Re-running is safe — it skips any university that's already in the DB.
"""

import re
import sqlite3
import os
import sys
import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency — run:  pip install openpyxl")
    sys.exit(1)

# ── Config ─────────────────────────────────────────────────────────────────────

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "tracker.db")

# Deadlines will be assumed to be in this year
# (you're applying for Fall 2028 entry → deadlines mostly early 2027)
DEADLINE_YEAR = 2027

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}

DEFAULT_CHECKLIST = [
    ("Statement of Purpose",         "SOP"),
    ("CV / Resume",                  "CV"),
    ("Official Transcripts",         "Transcript"),
    ("English Test (IELTS / TOEFL)", "Test Score"),
    ("Portfolio / Writing Sample",   "Portfolio"),
    ("Online Application Form",      "Application"),
    ("Application Fee Payment",      "Fee"),
]

# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_deadline(text, year=DEADLINE_YEAR):
    """
    Try to extract a date from a free-text timeline string.
    Looks for the word 'deadline' then scans the following text for a date.
    Returns a YYYY-MM-DD string or None.
    """
    if not text:
        return None
    t = str(text).lower()

    # Find everything after "deadline"
    m = re.search(r"deadline[s]?[:\s~]*(.{0,60})", t)
    chunk = m.group(1) if m else t

    # Non-EU / EU split — prefer non-EU (earlier) deadline
    non_eu = re.search(r"non[- ]eu[):]?\s*(\w+ \d{1,2}|\d{1,2} \w+|\w+)", chunk)
    if non_eu:
        chunk = non_eu.group(1)

    for name, num in MONTHS.items():
        # "Month DD" or "DD Month"
        p = re.search(rf"\b{name}\s+(\d{{1,2}})\b", chunk)
        if p:
            return f"{year}-{num:02d}-{int(p.group(1)):02d}"
        p = re.search(rf"\b(\d{{1,2}})\s+{name}\b", chunk)
        if p:
            return f"{year}-{num:02d}-{int(p.group(1)):02d}"
        # "late/early/mid Month" or bare "Month"
        p = re.search(rf"\b(late|early|mid)?\s*{name}\b", chunk)
        if p:
            qualifier = p.group(1)
            day = 25 if qualifier == "late" else 5 if qualifier == "early" else 15
            return f"{year}-{num:02d}-{day:02d}"

    return None


def map_priority(probability_text):
    """Map probability-of-admission text to reach / target / safety."""
    if not probability_text:
        return "target"
    p = str(probability_text).lower()
    if "medium-low" in p or "low-medium" in p:
        return "reach"
    if "low" in p:
        return "reach"
    if "medium-high" in p or "high" in p:
        return "target"   # still uncertain enough to call target, not safety
    return "target"


def extract_degree(program_name):
    """Infer degree type from program name."""
    if not program_name:
        return "MA"
    n = str(program_name)
    nu = n.upper()
    if "MFA" in nu:
        return "MFA"
    if "MSC" in nu or "M.SC" in nu or re.search(r"\bMS\b", nu):
        return "MSc"
    if re.search(r"\bMA\b", nu) or "M.A." in nu:
        return "MA"
    if "MDES" in nu:
        return "MA"
    return "MA"


def clean(val, maxlen=None):
    """Return a stripped string (or empty string for None/NaN)."""
    if val is None:
        return ""
    s = str(val).strip()
    if maxlen:
        s = s[:maxlen]
    return s


def portfolio_is_required(portfolio_text):
    """Return True if the portfolio field says it's required/needed."""
    if not portfolio_text:
        return False
    t = str(portfolio_text).lower()
    return any(w in t for w in ("yes", "required", "mandatory", "essential", "needed"))


# ── Main import ────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py path/to/your_file.xlsx")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print("Database not found. Run  python app.py  once to create it, then re-run this script.")
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row

    # Collect existing (university, program) pairs to avoid duplicates
    existing = set(
        (r["university"].strip().lower(), r["program"].strip().lower())
        for r in db.execute("SELECT university, program FROM applications").fetchall()
    )

    now = datetime.datetime.now().isoformat(timespec="seconds")
    added = 0
    skipped_dup = 0
    skipped_empty = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        country = sheet_name.strip()
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # First row is the header
        header = [clean(h) for h in rows[0]]

        def col(row_tuple, name):
            """Get value from row by column name."""
            try:
                idx = header.index(name)
                return clean(row_tuple[idx])
            except (ValueError, IndexError):
                return ""

        print(f"\n── {country} ({len(rows) - 1} programs) ──")

        for raw_row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in raw_row):
                skipped_empty += 1
                continue

            university   = col(raw_row, "Uni name")
            program      = col(raw_row, "Course name")
            timeline     = col(raw_row, "Application timeline")
            probability  = col(raw_row, "Probability of getting in")
            course_url   = col(raw_row, "Course websites")
            tuition      = col(raw_row, "Tuition")
            scholarships = col(raw_row, "Scholarships if any")
            ielts        = col(raw_row, "IELTS Requirement")
            portfolio    = col(raw_row, "Portfolio Required")
            semester     = col(raw_row, "Semester offered")
            requirements = col(raw_row, "Application requirements")
            where_apply  = col(raw_row, "Where to apply")

            if not university or not program:
                skipped_empty += 1
                continue

            key = (university.lower(), program.lower())
            if key in existing:
                print(f"  SKIP (already exists): {university}")
                skipped_dup += 1
                continue

            app_deadline   = parse_deadline(timeline)
            priority       = map_priority(probability)
            degree_type    = extract_degree(program)
            funding_type   = scholarships[:120] if scholarships else ""
            test_required  = ielts[:80] if ielts else ""
            needs_portfolio = portfolio_is_required(portfolio)

            # Build a useful notes block from the extra columns
            notes_parts = []
            if timeline:
                notes_parts.append(f"Timeline: {timeline}")
            if semester:
                notes_parts.append(f"Semester: {semester}")
            if requirements:
                notes_parts.append(f"Requirements: {requirements}")
            if tuition:
                notes_parts.append(f"Tuition: {tuition}")
            if where_apply:
                notes_parts.append(f"Apply via: {where_apply}")
            if probability:
                notes_parts.append(f"Admission chance: {probability}")
            notes = "\n".join(notes_parts)

            cur = db.execute("""
                INSERT INTO applications
                  (university, program, degree_type, country, funding_type,
                   application_deadline, status, priority,
                   test_required, application_url, notes, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                university, program, degree_type, country, funding_type,
                app_deadline, "researching", priority,
                test_required, course_url, notes, now, now,
            ))
            app_id = cur.lastrowid

            # Default checklist
            for i, (title, category) in enumerate(DEFAULT_CHECKLIST):
                db.execute("""
                    INSERT INTO checklist_items
                      (application_id, title, category, status, sort_order, created_at)
                    VALUES (?,?,?,?,?,?)
                """, (app_id, title, category, "not_started", i, now))

            # If portfolio is required, add a note to the portfolio checklist item
            if needs_portfolio:
                db.execute("""
                    UPDATE checklist_items SET notes = ?
                    WHERE application_id = ? AND category = 'Portfolio'
                """, (f"Required — {portfolio}", app_id))

            existing.add(key)
            added += 1
            deadline_str = f" (deadline ~{app_deadline})" if app_deadline else ""
            print(f"  ✓ {university} — {program}{deadline_str}")

    db.commit()
    db.close()

    print(f"""
────────────────────────────────
Import complete
  Added:   {added} applications
  Skipped: {skipped_dup} (already in DB)
  Blank:   {skipped_empty} (empty rows)
────────────────────────────────
Open http://localhost:5001 to see them.
""")


if __name__ == "__main__":
    main()
